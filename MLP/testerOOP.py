import math
from typing import Any
import random
import openpyxl

numberOfHiddenNodes = 4
sizeOfHiddenLayer = 4
numberOfPredictors = 8
numberOfOutputNodes = 1
indexOfOutputNode = -1

initialBias = 0.1
initialWeights = 0.1
learningParameter = 0.01
randomParameterLowerLimit = -2/numberOfPredictors
randomParameterUpperLimit = 2/numberOfPredictors

numberOfParameters = sizeOfHiddenLayer * 8 + sizeOfHiddenLayer + sizeOfHiddenLayer + 1

startingWeightList = [""]
startingBiasList = [""]
actualList = []

errorList = []
weightList = []
nodeList = []


class InputNode():
    def __init__(self, number, input):
        self.number = number
        self.uValue = input

    def updateInputValue(self, newInput):
        self.uValue = newInput

class Node:
    def __init__(self, number):
        self.number = number 
        self.bias = random.uniform(randomParameterLowerLimit, randomParameterUpperLimit)
        startingBiasList.append(self.bias)
        self.weightedSum = 0
        self.uValue = 0
        self.delta = 0

    def updateWeightedSum(self):
        self.weightedSum = self.bias
        for node in nodeList:
            for weight in weightList:
                if (weight.startNode == node.number) and (weight.endNode == self.number):
                    self.weightedSum += weight.weightValue * node.uValue
    
    def updateUValue(self):
        self.uValue = 1 / (1 + math.exp(-self.weightedSum))

    def updateBias(self):
        self.bias = self.bias + learningParameter * self.delta

class HiddenNode(Node):
    def __init__(self, number):
        super().__init__(number + numberOfPredictors + numberOfOutputNodes)
        for x in range (1, numberOfPredictors):
            weightList.append(Weight(x, self.number))

    def updateDelta(self):
        fDash = self.uValue * (1 - self.uValue)
        for node in nodeList:
            if node.number == 0:
                delta0 = node.delta
        for weight in weightList:
            if (weight.startNode == self.number) and (weight.endNode == 0):
                weightValue = weight.weightValue
                self.delta = weightValue * delta0 * fDash
                break
    
class OutputNode(Node):
    def __init__(self, actualValue):
        super().__init__(0)
        self.actualValue = actualValue
        actualList.append(actualValue)
        for x in range (1, numberOfHiddenNodes):
            weightList.append(Weight(x, self.number))

    def updateActualValue(self, newActual):
        self.actualValue = newActual

    def updateDelta(self):
        fDash = self.uValue * (1 - self.uValue)
        self.delta = (self.actualValue - self.uValue) * fDash
    
class Weight():
    def __init__(self, startNode, endNode):
        self.startNode, self.endNode = startNode, endNode
        self.weightValue = random.uniform(randomParameterLowerLimit, randomParameterUpperLimit)
        startingWeightList.append(self.weightValue)

    def updateWeights(self):
        oldWeight = self.weightValue
        for node in nodeList:
            if node.number == self.endNode:
                endDelta = node.delta
            if node.number == self.startNode:
                startValue = node.uValue
        self.weightValue = oldWeight + learningParameter * endDelta * startValue

def Initialize(entry):
    nodeList.append(OutputNode(entry[-1]))
    for x in range(1, numberOfPredictors + 1):
        nodeList.append(InputNode(x, entry[x]))
    for x in range(0, numberOfHiddenNodes):
        nodeList.append(HiddenNode(x))

def Reinitialize(entry):
    nodeList[0].updateActualValue(entry[indexOfOutputNode])

    for x in range(1, numberOfPredictors):
        nodeList[x].updateInputValue(entry[x])

def ForwardPass(modelledList):
    for node in nodeList:
        if isinstance(node, HiddenNode):
            node.updateWeightedSum()
            node.updateUValue()

    for node in nodeList:
        if isinstance(node, OutputNode):
            node.updateWeightedSum()
            node.updateUValue()
            modelledList.append(node.uValue)

def BackWardPass():
    for node in nodeList:
        if isinstance(node, OutputNode):
            node.updateDelta()

    for node in nodeList:
        if isinstance(node, HiddenNode):
            node.updateDelta()

def WeightAndBiasUpdate():
    for weight in weightList:
        weight.updateWeights()

    for node in nodeList:
        if isinstance(node, HiddenNode):
            node.updateBias()

    for node in nodeList:
        if isinstance(node, OutputNode):
            node.updateBias()

def RMSE(modelledResults, actualResults):
    squared_diffs = [(p - t) ** 2 for p, t in zip(modelledResults, actualResults)]
    mean_squared_error = sum(squared_diffs) / len(modelledResults)
    rmse = math.sqrt(mean_squared_error)
    return rmse

def main1():
    read_path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions->Sieved->Split->Standard.xlsx"
 
    read_wb_obj = openpyxl.load_workbook(read_path, data_only=True)
    
    sheet_obj = read_wb_obj['StandardCalibrate']
    
    rowMax = sheet_obj.max_row
    columnMax = sheet_obj.max_column

    entryList = []

    for i in range(3, rowMax + 1):
        entry = []
        for j in range(1, columnMax + 1):
            entry.append(sheet_obj.cell(row = i, column = j).value)
        entryList.append(entry)

    for q in range(4, 17):
        print(q)

        errorList.clear()
        weightList.clear()
        nodeList.clear()

        global numberOfHiddenNodes
        numberOfHiddenNodes = q
        Initialize(entryList[0])
        for x in range(0, 100):
            modelledList = []
            for entry in entryList:
                ForwardPass(modelledList)
                BackWardPass()
                WeightAndBiasUpdate()
                try:
                    Reinitialize(entry)
                except Exception as e:
                    break
            errorList.append(RMSE(modelledList, actualList))
            modelledList = []

        write_path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/MLP/MLPTraining.xlsx"
    
        write_wb_obj = openpyxl.load_workbook(write_path, data_only=True)
        
        new_sheet_obj = write_wb_obj["HiddenNodes"]

        for y in range (1, len(errorList) + 1):
            new_sheet_obj.cell(row = q, column = y).value = errorList[y - 1]

        write_wb_obj.save(write_path)

def main():
    read_path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions->Sieved->Split->Standard.xlsx"
 
    read_wb_obj = openpyxl.load_workbook(read_path, data_only=True)
    
    sheet_obj = read_wb_obj['StandardCalibrate']
    
    rowMax = sheet_obj.max_row
    columnMax = sheet_obj.max_column

    entryList = []

    for i in range(3, rowMax + 1):
        entry = []
        for j in range(1, columnMax + 1):
            entry.append(sheet_obj.cell(row = i, column = j).value)
        entryList.append(entry)

    newSheetData = []

    Initialize(8, entryList[0])
    for x in range(0, 100):
        modelledList = []
        for entry in entryList:
            ForwardPass(modelledList)
            BackWardPass()
            WeightAndBiasUpdate()
            try:
                Reinitialize(entry)
            except Exception as e:
                break
        errorList.append(RMSE(modelledList, actualList))
        modelledList = []

    for each in errorList:
        print(each)

main1()


