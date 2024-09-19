import math
from typing import Any
import random
import openpyxl
import numpy as np

numberOfHiddenNodes = 4
sizeOfHiddenLayer = 4
numberOfPredictors = 8
numberOfOutputNodes = 1
indexOfOutputNode = -1

initialBias = 0.1
initialWeights = 0.1
learningParameter = 0.1

#Momentum
global momentumConstant
momentumConstant = 0.1

randomParameterLowerLimit = -2/numberOfPredictors
randomParameterUpperLimit = 2/numberOfPredictors

numberOfParameters = sizeOfHiddenLayer * 8 + sizeOfHiddenLayer + sizeOfHiddenLayer + 1

startingWeightList = [0.007481429,-0.006055067,0.194138894,0.063997703,-0.24212507,0.094599025,0.229844603,0.175924805,-0.058898191,0.153361863,0.133685957,0.131435074,0.1715488,0.119481675,-0.096104571,0.008283301,-0.231386733,-0.174494865,0.011758528,0.109356964,0.112874608,0.104963345,0.116422118,0.14664802,-0.167459264,0.104601404,0.060889111,-0.185747669,-0.215454776,0.139987876,-0.141268565,0.192510289,0.091499703,-0.001328184,0.228226443,-0.00235409]
startingBiasList = [0.224315818, 0.100207781, 0.138279576, 0.072573976, -0.139837395]
actualList = []

errorList = []
weightList = []
nodeList = []


class InputNode():
    def __init__(self, number, input):
        self.number = number
        self.uValue = input

    def updateInputValue(self, newInput):
        self.uValue = newInput

class Node:
    def __init__(self, number, bias):
        self.number = number 
        self.bias = bias
        self.weightedSum = 0
        self.uValue = 0
        self.delta = 0

    def updateWeightedSum(self):
        self.weightedSum = self.bias
        for node in nodeList:
            for weight in weightList:
                if (weight.startNode == node.number) and (weight.endNode == self.number):
                    self.weightedSum += weight.weightValue * node.uValue
    
    def updateUValue(self):
        self.uValue = 1 / (1 + np.exp(-self.weightedSum))

    def updateBias(self):
        self.bias = self.bias + learningParameter * self.delta

class HiddenNode(Node):
    def __init__(self, number, bias):
        super().__init__(number + numberOfPredictors, bias)
        for x in range (1, numberOfPredictors + 1):
            weightList.append(Weight(x, self.number, startingWeightList[(self.number - numberOfPredictors) * numberOfHiddenNodes + x - 1]))

    def updateDelta(self):
        fDash = self.uValue * (1 - self.uValue)
        for node in nodeList:
            if node.number == 0:
                delta0 = node.delta
        for weight in weightList:
            if (weight.startNode == self.number) and (weight.endNode == 0):
                weightValue = weight.weightValue
                self.delta = weightValue * delta0 * fDash
                break
    
class OutputNode(Node):
    def __init__(self, actualValue, bias):
        super().__init__(0, bias)
        self.actualValue = actualValue
        actualList.append(actualValue)
        for x in range (1, numberOfHiddenNodes):
            weightList.append(Weight(x, self.number, startingWeightList[x]))

    def updateActualValue(self, newActual):
        self.actualValue = newActual

    def updateDelta(self):
        fDash = self.uValue * (1 - self.uValue)
        self.delta = (self.actualValue - self.uValue) * fDash
    
class Weight():
    def __init__(self, startNode, endNode, weightValue):
        self.startNode, self.endNode = startNode, endNode
        self.weightValue = weightValue
        self.momentumValue = 0

    def updateWeights(self):
        oldWeight = self.weightValue
        for node in nodeList:
            if node.number == self.endNode:
                endDelta = node.delta
            if node.number == self.startNode:
                startValue = node.uValue
        weightChange = learningParameter * endDelta * startValue + momentumConstant * self.momentumValue
        self.weightValue += weightChange
        self.momentumValue = weightChange


def Initialize(numberOfHiddenNodes, entry):
    for x in range(0, numberOfHiddenNodes + 1):
        startingBiasList.append(random.uniform(randomParameterLowerLimit, randomParameterUpperLimit))
    for x in range(0, numberOfPredictors * numberOfHiddenNodes + numberOfHiddenNodes):
        startingWeightList.append(random.uniform(randomParameterLowerLimit, randomParameterUpperLimit))
    
    
    nodeList.append(OutputNode(entry[indexOfOutputNode], startingBiasList[0]))
    for x in range(1, numberOfPredictors + 1):
        nodeList.append(InputNode(x, entry[x]))
    for x in range(1, numberOfHiddenNodes + 1):
        nodeList.append(HiddenNode(x, startingBiasList[x]))

def Reconstruct(numberOfHiddenNodes, entry):
    nodeList.append(OutputNode(entry[indexOfOutputNode], startingBiasList[0]))
    for x in range(1, numberOfPredictors + 1):
        nodeList.append(InputNode(x, entry[x]))
    for x in range(1, numberOfHiddenNodes + 1):
        nodeList.append(HiddenNode(x, startingBiasList[x]))

def Reinitialize(entry):
    nodeList[0].updateActualValue(entry[indexOfOutputNode])

    for x in range(1, numberOfPredictors):
        nodeList[x].updateInputValue(entry[x])

def ForwardPass(modelledList):
    for node in nodeList:
        if isinstance(node, HiddenNode):
            node.updateWeightedSum()
            node.updateUValue()

    for node in nodeList:
        if isinstance(node, OutputNode):
            node.updateWeightedSum()
            node.updateUValue()
            modelledList.append(node.uValue)

def BackWardPass():
    for node in nodeList:
        if isinstance(node, OutputNode):
            node.updateDelta()

    for node in nodeList:
        if isinstance(node, HiddenNode):
            node.updateDelta()

def WeightAndBiasUpdate():
    for weight in weightList:
        weight.updateWeights()

    for node in nodeList:
        if isinstance(node, HiddenNode):
            node.updateBias()

    for node in nodeList:
        if isinstance(node, OutputNode):
            node.updateBias()

def RMSE(modelledResults, actualResults):
    squared_diffs = [(p - t) ** 2 for p, t in zip(modelledResults, actualResults)]
    mean_squared_error = sum(squared_diffs) / len(modelledResults)
    rmse = math.sqrt(mean_squared_error)
    return rmse

def main1():
    read_path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions->Pure->Trimmed->Standard.xlsx"
 
    read_wb_obj = openpyxl.load_workbook(read_path, data_only=True)
    
    sheet_obj = read_wb_obj['StandardCalibrate']
    
    rowMax = sheet_obj.max_row
    columnMax = sheet_obj.max_column

    entryList = []

    for i in range(3, rowMax + 1):
        entry = []
        for j in range(1, columnMax + 1):
            entry.append(sheet_obj.cell(row = i, column = j).value)
        entryList.append(entry)

    q = 9000
    count = 1
    while count < 11:
        

        errorList.clear()
        errorList.append(q)
        weightList.clear()
        nodeList.clear()

        global momentumConstant
        momentumConstant = q
        print(momentumConstant)
        Reconstruct(4, entryList[0])
        for x in range(0, 100):
            modelledList = []
            for entry in entryList:
                ForwardPass(modelledList)
                BackWardPass()
                WeightAndBiasUpdate()
                try:
                    Reinitialize(entry)
                except Exception as e:
                    break
            errorList.append(RMSE(modelledList, actualList))
            modelledList = []

        write_path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/MLP/MLPTraining.xlsx"
    
        write_wb_obj = openpyxl.load_workbook(write_path, data_only=True)
        
        new_sheet_obj = write_wb_obj["MomentumOutliers"]

        for y in range (1, len(errorList) + 1):
            new_sheet_obj.cell(row = count, column = y).value = errorList[y - 1]

        count += 1
        q = q/10

        write_wb_obj.save(write_path)

def main():
    read_path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions->Sieved->Split->Standard.xlsx"
 
    read_wb_obj = openpyxl.load_workbook(read_path, data_only=True)
    
    sheet_obj = read_wb_obj['StandardCalibrate']
    
    rowMax = sheet_obj.max_row
    columnMax = sheet_obj.max_column

    entryList = []

    for i in range(3, rowMax + 1):
        entry = []
        for j in range(1, columnMax + 1):
            entry.append(sheet_obj.cell(row = i, column = j).value)
        entryList.append(entry)

    newSheetData = []

    Initialize(5, entryList[0])
    for x in range(0, 100):
        modelledList = []
        for entry in entryList:
            ForwardPass(modelledList)
            BackWardPass()
            WeightAndBiasUpdate()
            try:
                Reinitialize(entry)
            except Exception as e:
                break
        errorList.append(RMSE(modelledList, actualList))
        modelledList = []

    for each in errorList:
        print(each)

main1()


