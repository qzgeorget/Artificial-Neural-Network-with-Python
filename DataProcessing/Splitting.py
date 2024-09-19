import openpyxl

path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions->Pure->Trimmed->Standard.xlsx"
 
wb_obj = openpyxl.load_workbook(path, data_only=True)
 
sheet_obj = wb_obj['Trimmed']
calibrate_sheet_obj = wb_obj.create_sheet(title='CalibrateSet')
validate_sheet_obj = wb_obj.create_sheet(title='ValidateSet')
test_sheet_obj = wb_obj.create_sheet(title='TestSet')

rowMax = sheet_obj.max_row
columnMax = sheet_obj.max_column + 1

ci = 1
vi = 1
ti = 1

for x in range(1, rowMax, 5):
    for i in range(0, 3): 
        for y in range (1, columnMax):
            calibrate_sheet_obj.cell(row=ci,column=y).value = sheet_obj.cell(row=x + i,column=y).value
        ci += 1
    for y in range (1, columnMax):
        validate_sheet_obj.cell(row=vi,column=y).value = sheet_obj.cell(row=x + 3,column=y).value
    for y in range (1, columnMax):
        test_sheet_obj.cell(row=ti,column=y).value = sheet_obj.cell(row=x + 4,column=y).value
    vi += 1
    ti += 1



wb_obj.save(path)
    