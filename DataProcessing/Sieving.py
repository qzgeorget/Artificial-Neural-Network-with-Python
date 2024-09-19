import openpyxl
from openpyxl.styles import PatternFill, Border, Side
 
# Give the location of the file
path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions.xlsx"
 
wb_obj = openpyxl.load_workbook(path, data_only=True)
 
sheet_obj = wb_obj['Pure']
organised_sheet_obj = wb_obj.create_sheet(title='Sieved')
 
rowMax = sheet_obj.max_row
columnMax = sheet_obj.max_column

sdArray = []
for i in range(1, columnMax + 1):
    sdArray.append(sheet_obj.cell(row = 2, column = i).value)

averageArray = []
for i in range(1, columnMax + 1):
    averageArray.append(sheet_obj.cell(row = 3, column = i).value)


usableEntries = 0
x = 1
for i in range(4, rowMax + 1):
    emptyFlag = False
    outlierFlag = False
    for j in range(1, columnMax + 1):
        cell_val = sheet_obj.cell(row=i, column=j).value

        if not (isinstance(cell_val, float) or isinstance(cell_val, int)) or (cell_val == -999):
            emptyFlag = True

    if (emptyFlag == True):
        for j in range(1, columnMax + 1):
            sheet_obj.cell(row=i, column=j).value = None
    else: 
        usableEntries += 1
        for j in range(1, columnMax + 1):
            cell_val = sheet_obj.cell(row=i,column=j).value
            organised_sheet_obj.cell(row=x,column=j).value = cell_val
        x += 1


print("usableEntries: ", usableEntries)

wb_obj.save(path)


