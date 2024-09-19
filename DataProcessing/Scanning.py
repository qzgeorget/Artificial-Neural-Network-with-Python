import openpyxl
from openpyxl.styles import PatternFill, Border, Side
 
# Give the location of the file
path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions Pure.xlsx"
 
wb_obj = openpyxl.load_workbook(path, data_only=True)
 
sheet_obj = wb_obj['Pure']
 
rowMax = sheet_obj.max_row
columnMax = sheet_obj.max_column

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='lightUp')
blueFill = PatternFill(start_color='0000FFFF',
                   end_color='0000FFFF',
                   fill_type='lightUp')

thick_border = Border(
    left=Side(style='thick', color='000000'),
    right=Side(style='thick', color='000000'),
    top=Side(style='thick', color='000000'),
    bottom=Side(style='thick', color='000000')
)

sdArray = []
for i in range(1, columnMax + 1):
    sdArray.append(sheet_obj.cell(row = 2, column = i).value)

averageArray = []
for i in range(1, columnMax + 1):
    averageArray.append(sheet_obj.cell(row = 3, column = i).value)


usableEntries = 0
for i in range(4, rowMax + 1):
    emptyFlag = False
    outlierFlag = False
    for j in range(1, columnMax + 1):
        cell_obj = sheet_obj.cell(row=i, column=j)

        if not (isinstance(cell_obj.value, float) or isinstance(cell_obj.value, int)) or (cell_obj.value == -999):
            emptyFlag = True
            cell_obj.border = thick_border

        try: 
            if ((cell_obj.value > averageArray[j-1] + sdArray[j-1] * 2.5) or (cell_obj.value < averageArray[j-1] - sdArray[j-1] * 2.5)):
                outlierFlag = True
                cell_obj.border = thick_border
        except (TypeError):
            next

    if (emptyFlag == True):
        for j in range(1, columnMax + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            cell_obj.fill = redFill
    elif (outlierFlag == True):
        for j in range(1, columnMax + 1):
            cell_obj = sheet_obj.cell(row=i, column=j)
            cell_obj.fill = blueFill
    else:
        usableEntries += 1

print("usableEntries: ", usableEntries)

wb_obj.save(path)


