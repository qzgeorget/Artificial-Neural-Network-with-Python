import openpyxl
from openpyxl.styles import PatternFill, Border, Side
 
# Give the location of the file
path = "/Users/driplord3000/Documents/Loughborough/Part B/Semester 2/AI Methods/Sem2Coursework/ANN Code/DataProcessing/PythonInteractions->Sieved->Split.xlsx"
 
wb_obj = openpyxl.load_workbook(path, data_only=True)
 
sheet_obj = wb_obj['TestSet']
standard_sheet_obj = wb_obj.create_sheet(title='StandardTest')
 
rowMax = sheet_obj.max_row
columnMax = sheet_obj.max_column

maxArray = []
for i in range(1, columnMax + 1):
    maxArray.append(sheet_obj.cell(row = 1, column = i).value)


minArray = []
for i in range(1, columnMax + 1):
    minArray.append(sheet_obj.cell(row = 2, column = i).value)

for j in range(1, columnMax + 1):
    columnMax = maxArray[j-1]
    columnMin = minArray[j-1]
    columnRange = columnMax-columnMin
    for i in range(3, rowMax + 1):
        cell_val = sheet_obj.cell(row = i, column = j).value
        newValue = 0.1 + 0.8 * ((cell_val - columnMin)/columnRange)
        standard_sheet_obj.cell(row = i, column = j).value = newValue


    


wb_obj.save(path)


